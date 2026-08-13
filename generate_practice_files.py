"""
生成数据分析打卡网站的配套练习Excel文件
每个教程日一个文件，使用零售/电商主题数据
"""
import os
import random
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 固定随机种子: 每次生成的数据一致,方便对照练习答案
random.seed(42)

OUTPUT_DIR = os.path.join(os.path.dirname(__file__), 'practice-files')

# 通用样式
HEADER_FONT = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center')
DATA_FONT = Font(name='微软雅黑', size=10)
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)

def style_header(ws, headers, col_widths=None):
    """给工作表添加表头样式"""
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
    if col_widths:
        for col_idx, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = w

def style_data(ws, start_row, end_row, max_col):
    """给数据区域加样式"""
    for row in range(start_row, end_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER

def add_instruction_sheet(wb, title, instructions, thinking=None):
    """添加练习说明工作表,可在末尾追加折叠答案的思考题"""
    ws = wb.create_sheet('练习说明', 0)
    ws.column_dimensions['A'].width = 80
    ws.cell(row=1, column=1, value=title).font = Font(name='微软雅黑', bold=True, size=14)
    for i, line in enumerate(instructions, 3):
        ws.cell(row=i, column=1, value=line).font = Font(name='微软雅黑', size=10.5)
        ws.cell(row=i, column=1).alignment = Alignment(wrap_text=True)
    # 删除openpyxl新建工作簿时自带的空工作表
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    if thinking:
        add_thinking(ws, 3 + len(instructions) + 1, thinking)
    return ws


def add_thinking(ws, start_row, questions):
    """在指定工作表追加思考题: 答案行默认折叠,点左侧+号展开查看"""
    r = start_row
    c = ws.cell(row=r, column=1, value='思考题（先自己想想，点左侧+号展开答案）')
    c.font = Font(name='微软雅黑', bold=True, size=11, color='E67E22')
    r += 1
    for qi, (q, a) in enumerate(questions, 1):
        ws.cell(row=r, column=1, value=f'思考题{qi}: {q}').font = Font(name='微软雅黑', bold=True, size=10.5)
        r += 1
        ans_start = r
        for j, line in enumerate(a.split('\n')):
            c = ws.cell(row=r, column=1, value=('答案: ' if j == 0 else '') + line)
            c.font = Font(name='微软雅黑', size=10.5, color='2E7D32')
            c.alignment = Alignment(wrap_text=True)
            r += 1
        # 答案行折叠(Excel/WPS都支持行分组,点左侧+号展开)
        ws.row_dimensions.group(ans_start, r - 1, outline_level=1, hidden=True)
        r += 1
    return r


def add_thinking_sheet(wb, questions):
    """独立"思考题"工作表(用于没有练习说明表的文件)"""
    ws = wb.create_sheet('思考题')
    ws.column_dimensions['A'].width = 80
    ws.cell(row=1, column=1, value='思考题（先自己想想，点左侧+号展开答案）').font = Font(name='微软雅黑', bold=True, size=13, color='E67E22')
    for qi, (q, a) in enumerate(questions, 1):
        ws.cell(row=2 + (qi - 1) * 4, column=1, value=f'思考题{qi}: {q}').font = Font(name='微软雅黑', bold=True, size=11)
        ans_start = 3 + (qi - 1) * 4
        for j, line in enumerate(a.split('\n')):
            c = ws.cell(row=ans_start + j, column=1, value=('答案: ' if j == 0 else '') + line)
            c.font = Font(name='微软雅黑', size=10.5, color='2E7D32')
            c.alignment = Alignment(wrap_text=True)
        ws.row_dimensions.group(ans_start, ans_start + len(a.split('\n')) - 1, outline_level=1, hidden=True)
    return ws


# ==================== 各文件思考题(答案折叠) ====================
DAY02_THINKING = [
    ('文本"123"和数字123在筛选时有什么不同？怎么一眼分辨？', '筛选"大于100"时只有数字123会被筛出来，文本"123"不参与数值比较。一眼分辨看对齐方式：文本左对齐、数字右对齐（左上角有绿色小三角提示的也是文本）。'),
    ('Ctrl+E（智能填充）什么时候会失灵？补完后必须做什么？', '数据没有统一规律时，补出来的结果是错的。补完之后必须从上到下扫一眼全部结果，不对的单元格手动改。'),
]
DAY03_THINKING = [
    ('筛选和排序的本质区别？发报告给同事前应该用哪个？', '筛选只是"隐藏"不符合条件的行，数据还在原位；排序是真正"重新排列"行的顺序。发报告前先排好序再发，别人打开一眼能看懂。'),
    ('多条件排序"先按位数升序、再按价格降序"是怎么排的？', '先按位数分组排，位数相同的行内部再按价格从大到小排。是"先主后次"的关系，不是两个条件混成一次排序。'),
]
DAY04_THINKING = [
    ('去重时勾一列和勾两列的区别？给发奖名单去重应该勾什么？', '勾一列=该列内容重复就删（同一用户的多条记录会被误删）；勾两列=两列同时相同才视为重复。发奖名单按"用户ID"单列去重。'),
    ('为什么去重和替换之前一定要先复制一份原始数据？', '这两个操作不可逆，发奖励场景删错了是真金白银的损失。"动刀前先备份"是数据分析的第一条职业习惯。'),
]
DAY05_THINKING = [
    ('数据验证和条件格式都跟"数据质量"有关，区别是什么？', '数据验证是"预防"——在数据进来之前限制输入，不让错数据出现；条件格式是"体检"——数据已经存在，标出来让你看。一个管入口，一个管事后。'),
    ('本文件里价格>1000的商品有哪些？<100的又有哪些？这反映什么定价策略？', '>1000：iPhone 15 Pro(1299)、MacBook Air(2499)、空气净化器(1999)、扫地机器人(1299)、蚕丝被(1599)——集中在电子和家居。\n<100：机械键盘(89)、纯棉T恤(39)、台灯LED(89)、蓝牙音箱(79)、瑜伽垫(49)——以服装和户外为主。\n策略：电子/家居走高单价利润款，服装/户外走低价走量款。'),
]
DAY06_THINKING = [
    ('透视表四个区域，一句话分别是什么？', '行/列=按什么分类；值=汇总计算什么（计数/求和/平均）；筛选器=整体按什么条件过滤。'),
    ('透视表默认对数字字段和文本字段分别做什么汇总？想要"平均值"怎么办？', '数字默认求和、文本默认计数。要平均值：点值字段→值字段设置→汇总方式选"平均值"。'),
]
DAY07_THINKING = [
    ('什么时候需要日期分组？', '行标签是日期时每天一行几百行没法看。右键日期→组合→按月/季度/年归并，几百行变十几行，趋势一目了然。'),
    ('切片器比筛选下拉框好在哪？（至少2点）', '①一眼能看出当前筛的是什么；②可以同时联动多个透视表和图表；③给老板展示时点按钮更直观专业。'),
]
DAY09_THINKING = [
    ('COUNT和COUNTA混用会有什么后果？', 'COUNT只数数字，COUNTA数所有非空。给"用户名"这种文本列用COUNT会得到0，人数统计全错。统计前先想清楚列的类型。'),
    ('用AVERAGEIF算一下：本数据里哪个分类的平均客单价最高？', '答案是"电子产品"（平均1562.72元/单）。\n方法：=AVERAGEIF(分类列,某分类,金额列)，每个分类各算一遍再比较。'),
]
DAY10_THINKING = [
    ('VLOOKUP第四参数TRUE和FALSE各适合什么场景？', 'FALSE=精确匹配，找ID必须用它（找不到返回#N/A）；TRUE=近似匹配，"找最接近且不超过"，适合金额→折扣档位，但要求查找列升序。'),
    ('两张表都有"用户ID"，VLOOKUP却返回#N/A，最可能的原因？', '格式不一致——一边是文本一边是数字（看对齐方式分辨），或有多余空格。先统一两边格式再匹配。'),
]
DAY11_THINKING = [
    ('嵌套IF和IFS哪个更推荐？为什么？', 'IFS（Excel 2016+/WPS新版都有）。条件多时嵌套IF要一层层数括号，IFS平铺直叙，别人一眼能看懂判断逻辑。'),
    ('=TEXT(123,"000")的结果是数字123吗？', '不是，是文本"123"。TEXT返回的是文本，不能参与计算。需要计算时用原值。'),
]
DAY12_THINKING = [
    ('折线图和柱状图各适合什么场景？', '折线图=展示随时间的变化趋势；柱状图=分类之间比大小。分类数据画成折线图会暗示"连续性"，误导读者。'),
    ('饼图分类超过几个就不该用？', '5个。分类太多饼图碎成一片，读者分不清谁大谁小，改用柱状图。'),
]
DAY19_THINKING = [
    ('在本文件的模拟world表上，写出"查询亚洲国家名字和人口"的SQL。', "SELECT name, population FROM [表] WHERE continent = 'Asia'"),
    ('"WHERE population > 100000000 AND area > 1000000"会返回哪些国家？', 'China、India、United States、Indonesia、Brazil、Russia、Mexico、Egypt 共8个。注意：Japan、Nigeria人口达标但面积不够；Vietnam人口差一点不到1亿。'),
]
SQL_THINKING = [
    ('WHERE和HAVING的执行顺序？为什么"WHERE COUNT(*)>10"会报错？', '执行顺序：FROM→WHERE→GROUP BY→HAVING→SELECT→ORDER BY→LIMIT。WHERE在分组前执行，那时COUNT(*)还没算出来，所以分组后的条件只能放HAVING里。'),
    ('LEFT JOIN和INNER JOIN的区别？工作中哪个更常用？', 'LEFT JOIN保留左表所有行，右表没匹配就填NULL（"所有用户，买没买过都保留"）；INNER JOIN只留两边都匹配的行。工作中90%用LEFT JOIN。'),
    ('窗口函数和GROUP BY最大的区别？', 'GROUP BY把组内行合并成一行（原始明细丢失）；窗口函数保留所有原始行，只在每行旁边加一列计算结果（如组内排名）。'),
]

def add_data_sheet(wb, name, headers, rows, widths=None):
    """追加一个数据工作表(表头+样式)"""
    ws = wb.create_sheet(name)
    style_header(ws, headers, widths or [14] * len(headers))
    for i, row in enumerate(rows, 2):
        for j, val in enumerate(row, 1):
            ws.cell(row=i, column=j, value=val)
    style_data(ws, 2, len(rows) + 1, len(headers))
    return ws


def generate_day02():
    """Day 2: 数据输入与格式"""
    wb = Workbook()
    add_instruction_sheet(wb, 'Day 2 练习：数据输入与格式', [
        '本文件包含两个练习工作表：',
        '',
        '【数据类型练习】',
        '1. 观察A/B/C三列的对齐方式——文本左对齐？数字右对齐？',
        '2. 选中B列的日期，右键→设置单元格格式→分别切换到不同格式，看显示变化',
        '3. 在空白行自己输入不同类型的数据试试',
        '',
        '【Ctrl+E练习】',
        'WPS提示：WPS里叫"智能填充"，快捷键一样是Ctrl+E（需WPS 2019及以上版本）',
        '1. A列是"姓名+手机号"的合并数据',
        '2. 在B1手打第一行的名字（张三），B2手打第二行的名字（李四）',
        '3. 选中B1:B2→按Ctrl+E→Excel自动补全所有名字',
        '4. 同样方法在C列提取手机号',
        '5. 试试反过来：在D列把名字和手机号用"-"重新合并（如"张三-13800138001"）',
    ], thinking=DAY02_THINKING)

    # Sheet: 数据类型练习
    ws1 = wb.create_sheet('数据类型练习')
    headers = ['文本型数据', '数字型数据', '日期型数据']
    style_header(ws1, headers, [20, 18, 18])

    data = [
        ["'00123（文本）", 123, '2026-01-15'],
        ["'00456（文本）", 456, '2026-02-20'],
        ["'00789（文本）", 789.5, '2026-03-10'],
        ["'员工编号A001", 1000, '2026-04-05'],
        ["'商品SKU-B123", 2500.99, '2026-05-18'],
        ["'订单NO-2026", -150, '2026-06-22'],
        ["'用户ID:8888", 0, '2026-07-01'],
        ["'备注：已发货", 8888.88, '2026-08-06'],
    ]
    for i, row_data in enumerate(data, 2):
        for j, val in enumerate(row_data, 1):
            ws1.cell(row=i, column=j, value=val)
    ws1.cell(row=10, column=1, value="← 试试在这里输入不同类型的数据")
    style_data(ws1, 2, 10, 3)

    # Sheet: Ctrl+E练习
    ws2 = wb.create_sheet('Ctrl+E练习')
    headers2 = ['姓名+手机号（合并列）', '姓名（Ctrl+E拆分）', '手机号（Ctrl+E拆分）', '重新合并（Ctrl+E）']
    style_header(ws2, headers2, [28, 22, 22, 30])

    names_phones = [
        '张三13800138001', '李四13900139002', '王五13700137003',
        '赵六13600136004', '孙七13500135005', '周八13400134006',
        '吴九13300133007', '郑十13200132008', '钱十一13100131009',
        '冯十二15000150010', '陈十三15100151011', '褚十四15200152012',
        '卫十五15300153013', '蒋十六15500155014', '沈十七15600156015',
    ]
    for i, val in enumerate(names_phones, 2):
        ws2.cell(row=i, column=1, value=val)
    # B和C列留给用户练习，D列给提示
    ws2.cell(row=2, column=4, value='提示：在B列和C列先用Ctrl+E拆分，再在D列试试合并')
    style_data(ws2, 2, len(names_phones) + 1, 4)

    path = os.path.join(OUTPUT_DIR, 'day02-数据输入与格式.xlsx')
    wb.save(path)
    print(f'  [OK] {os.path.basename(path)}')


def generate_day03():
    """Day 3: 筛选与排序"""
    wb = Workbook()
    add_instruction_sheet(wb, 'Day 3 练习：筛选与排序', [
        '这是一份电商订单明细表（50行），请完成以下练习：',
        '',
        '【筛选练习】',
        '1. 筛选出"电子产品"分类的所有订单',
        '2. 筛选出价格>500的订单',
        '3. 筛选出"北京"且金额>200的订单（叠加筛选）',
        '4. 筛出商品名称包含"耳机"的订单（文本筛选→包含）',
        '',
        '【排序练习】',
        '1. 按金额从大到小排序，金额最高的10单是什么？',
        '2. 按"城市→金额降序"排序（多条件排序）',
        '3. 按"分类→日期升序"排序',
        '',
        '【自由探索】',
        '试试筛选+排序的组合：先筛出"上海"的订单，再按金额降序排',
        '【备用实战】没有真实靓号数据时，用"靓号数据(备用实战)"工作表练习（结构与真实数据一致）',
    ], thinking=DAY03_THINKING)

    ws = wb.create_sheet('订单明细')
    headers = ['订单ID', '商品名称', '分类', '价格', '数量', '金额', '日期', '城市']
    style_header(ws, headers, [12, 20, 12, 10, 8, 12, 14, 10])

    categories = {
        '电子产品': ['iPhone 15', 'AirPods Pro', 'iPad Air', 'MacBook Pro', '华为Mate 60', '小米14', '机械键盘', '蓝牙音箱', '显示器27寸', '移动硬盘2T'],
        '服装鞋帽': ['运动跑鞋', '羽绒服', '牛仔裤', 'T恤纯棉', '冲锋衣', '卫衣连帽', '休闲皮鞋', '棒球帽', '围巾羊绒', '运动短裤'],
        '食品饮料': ['有机牛奶', '坚果礼盒', '绿茶龙井', '咖啡豆', '巧克力礼盒', '蜂蜜柚子茶', '进口红酒', '蛋白粉', '橄榄油', '即食燕窝'],
        '家居用品': ['乳胶枕头', '蚕丝被', '空气净化器', '台灯LED', '收纳箱', '四件套纯棉', '扫地机器人', '加湿器', '地毯', '毛巾浴巾'],
        '运动户外': ['瑜伽垫', '登山包', '帐篷双人', '跑步腰包', '游泳镜', '羽毛球拍', '自行车头盔', '滑雪手套', '防晒衣', '望远镜'],
    }
    cities = ['北京', '上海', '广州', '深圳', '杭州', '成都', '武汉', '南京']

    rows = []
    for i in range(1, 51):
        cat = random.choice(list(categories.keys()))
        product = random.choice(categories[cat])
        price = round(random.uniform(29, 2999), 2)
        qty = random.randint(1, 5)
        amount = round(price * qty, 2)
        date = datetime(2026, random.randint(1, 8), random.randint(1, 28))
        city = random.choice(cities)
        rows.append([f'ORD-{i:04d}', product, cat, price, qty, amount, date.strftime('%Y-%m-%d'), city])

    for i, row_data in enumerate(rows, 2):
        for j, val in enumerate(row_data, 1):
            ws.cell(row=i, column=j, value=val)
    style_data(ws, 2, len(rows) + 1, len(headers))

    add_data_sheet(wb, '靓号数据(备用实战)', ['靓号ID', '位数', '类别', '价格(乐元)', '购买方式', '赠送VIP'], generate_haoid_rows(60))

    path = os.path.join(OUTPUT_DIR, 'day03-筛选与排序.xlsx')
    wb.save(path)
    print(f'  [OK] {os.path.basename(path)}')


def generate_day04():
    """Day 4: 数据清洗"""
    wb = Workbook()
    add_instruction_sheet(wb, 'Day 4 练习：数据清洗', [
        '这份"脏数据"故意制造了各种常见问题，请逐一清洗：',
        '',
        '【练习1：删除重复项】',
        '数据→删除重复项→找出哪些行是重复的？分别试试单条件去重和多条件去重',
        'WPS提示：WPS里在"数据"选项卡的"重复项"组里找"删除重复项"，还可先试试"高亮重复项"',
        '',
        '【练习2：文本分列】',
        'A列"用户ID-日期"是合并的→数据→分列→按分隔符"-" 拆成两列',
        'WPS提示：WPS里同样在"数据"选项卡→分列，流程一样',
        '',
        '【练习3：查找替换】',
        '1. 把多余的空格删掉（C列有些名字前后有空格）',
        '2. D列有些单元格里有换行符（Alt+Enter产生的），用查找替换清理',
        '',
        '【备用实战】"活动参与名单(备用实战)"工作表模拟活动回复用户列表（含重复用户），练去重用',
        '【注意】操作前先把"脏数据"工作表复制一份（右键标签→移动或复制→建立副本），防止改坏了没法恢复。',
    ], thinking=DAY04_THINKING)

    ws = wb.create_sheet('脏数据')
    headers = ['用户ID-日期（需分列）', '用户ID（分列结果）', '用户姓名（有多余空格）', '备注（有换行符）']
    style_header(ws, headers, [28, 16, 22, 35])

    data = [
        ['U1001-20260315', '', ' 张三 ', '第一行\n备注内容'],
        ['U1002-20260316', '', '李  四', '正常备注'],
        ['U1001-20260315', '', ' 张三 ', '第一行\n备注内容'],  # 重复行
        ['U1003-20260317', '', '  王五  ', '有\n换行\n的备注'],
        ['U1004-20260318', '', '赵六', '正常'],
        ['U1003-20260317', '', '  王五  ', '有\n换行\n的备注'],  # 重复行
        ['U1005-20260319', '', ' 孙 七 ', '正常备注'],
        ['U1006-20260320', '', '  周八', '带\n多行\n内容'],
        ['U1007-20260321', '', ' 吴九', '正常'],
        ['U1008-20260322', '', '  郑十  ', '又有一个\n换行符'],
        ['U1005-20260319', '', ' 孙 七 ', '正常备注'],  # 重复行
        ['U1009-20260323', '', '钱十一', '正常'],
        ['U1010-20260324', '', '  冯十二  ', '末尾也有换行\n'],
    ]
    for i, row_data in enumerate(data, 2):
        for j, val in enumerate(row_data, 1):
            ws.cell(row=i, column=j, value=val)
    style_data(ws, 2, len(data) + 1, len(headers))

    add_data_sheet(wb, '活动参与名单(备用实战)', ['用户ID', '昵称', '参与活动', '参与日期'], generate_participant_rows(34))

    path = os.path.join(OUTPUT_DIR, 'day04-数据清洗.xlsx')
    wb.save(path)
    print(f'  [OK] {os.path.basename(path)}')


def generate_day05():
    """Day 5: 数据验证与条件格式"""
    wb = Workbook()
    add_instruction_sheet(wb, 'Day 5 练习：数据验证与条件格式', [
        '这是一个商品定价表，请完成以下练习：',
        '',
        '【数据验证】',
        'WPS提示：WPS里这个功能叫"有效性"（或数据有效性），位置一样在"数据"选项卡，找不到就用顶部搜索框搜"有效性"',
        '1. 选中"分类"列→数据→数据验证/有效性→序列→来源输入"电子产品,服装鞋帽,食品饮料,家居用品,运动户外"（英文逗号）',
        '   → 现在分类列只能从下拉列表里选，不能随便打字了',
        '2. 选中"价格"列→数据验证/有效性→小数→介于13到3000→防止输入异常价格',
        '3. 选中"库存"列→数据验证/有效性→整数→介于0到9999',
        '',
        '【条件格式】',
        'WPS提示：WPS里也叫"条件格式"，位置一样在"开始"选项卡，数据条/色阶/图标集都有',
        '1. 选中价格列→开始→条件格式→突出显示→大于→1000→设为红色',
        '2. 再加一条：小于100→设为绿色',
        '3. 选中库存列→条件格式→数据条→一眼看出库存高低',
        '4. 试试色阶：选中价格列→条件格式→色阶→绿-黄-红',
        '【备用实战】"靓号定价(备用实战)"工作表模拟靓号定价表，练数据验证和条件格式用',
        '',
        '【思考】价格>1000的和<100的商品，分别是什么类型？这反映了什么定价策略？（答案见下方折叠区）',
    ], thinking=DAY05_THINKING)

    ws = wb.create_sheet('商品定价')
    headers = ['商品名称', '分类', '价格', '库存']
    style_header(ws, headers, [22, 14, 12, 10])

    products = [
        ('iPhone 15 Pro', '电子产品', 1299, 45),
        ('AirPods Pro', '电子产品', 199, 230),
        ('MacBook Air', '电子产品', 2499, 18),
        ('小米14', '电子产品', 899, 120),
        ('机械键盘', '电子产品', 89, 300),
        ('运动跑鞋', '服装鞋帽', 599, 150),
        ('羽绒服', '服装鞋帽', 899, 80),
        ('纯棉T恤', '服装鞋帽', 39, 500),
        ('牛仔裤', '服装鞋帽', 299, 200),
        ('冲锋衣', '服装鞋帽', 799, 65),
        ('有机牛奶', '食品饮料', 12.9, 1000),
        ('坚果礼盒', '食品饮料', 168, 300),
        ('进口红酒', '食品饮料', 258, 120),
        ('蛋白粉', '食品饮料', 399, 90),
        ('即食燕窝', '食品饮料', 688, 45),
        ('乳胶枕头', '家居用品', 299, 180),
        ('空气净化器', '家居用品', 1999, 30),
        ('扫地机器人', '家居用品', 1299, 55),
        ('台灯LED', '家居用品', 89, 400),
        ('蚕丝被', '家居用品', 1599, 25),
        ('蓝牙音箱', '电子产品', 79, 350),
        ('登山包', '运动户外', 459, 110),
        ('瑜伽垫', '运动户外', 49, 600),
        ('帐篷双人', '运动户外', 899, 40),
        ('望远镜', '运动户外', 299, 85),
    ]
    for i, (name, cat, price, stock) in enumerate(products, 2):
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value=cat)
        ws.cell(row=i, column=3, value=price)
        ws.cell(row=i, column=4, value=stock)
    style_data(ws, 2, len(products) + 1, len(headers))

    add_data_sheet(wb, '靓号定价(备用实战)', ['靓号ID', '类别', '价格(乐元)', '购买方式'],
                   [[r[0], r[2], r[3], r[4]] for r in generate_haoid_rows(60)])

    path = os.path.join(OUTPUT_DIR, 'day05-数据验证与条件格式.xlsx')
    wb.save(path)
    print(f'  [OK] {os.path.basename(path)}')


def generate_day06():
    """Day 6-7: 透视表基础（Day7复用同一文件）"""
    wb = Workbook()
    add_instruction_sheet(wb, 'Day 6-7 练习：数据透视表', [
        '这是100+行销售明细数据。请完成以下练习：',
        '',
        '【Day 6：透视表基础】',
        '1. 插入透视表→把"分类"拖到行→"金额"拖到值→各分类销售额汇总',
        '2. 再加"数量"到值→看看各分类销量',
        '3. 把"城市"拖到筛选器→切换不同城市看数据变化',
        '4. 修改汇总方式：金额从"求和"改成"平均值"→各分类客单价',
        '',
        '【Day 7：进阶】',
        '1. 日期分组：把日期拖到行→右键→组合→按月分组→看月度趋势',
        '2. 数值分组：把金额拖到行→右键→组合→步长200→看金额分布',
        '3. 插入切片器：选"分类"和"城市"→实现一键切换视角',
        '   WPS提示：WPS的切片器入口不同——先点击选中透视表→顶部出现"数据透视表工具"→在"分析"选项卡里点"插入切片器"（WPS 2019以上版本才有）',
        '4. 试试：同一个透视表同时看"各城市×各分类"的交叉销售额（行=城市，列=分类，值=金额求和）',
    ], thinking=DAY06_THINKING)

    ws = wb.create_sheet('销售明细')
    headers = ['日期', '商品名称', '分类', '城市', '数量', '单价', '金额']
    style_header(ws, headers, [14, 20, 12, 8, 8, 10, 12])

    categories = {
        '电子产品': ['iPhone 15', 'AirPods Pro', 'iPad Air', 'MacBook Pro', '华为Mate 60', '小米14', '机械键盘', '蓝牙音箱', '显示器', '移动硬盘'],
        '服装鞋帽': ['运动跑鞋', '羽绒服', '牛仔裤', 'T恤纯棉', '冲锋衣', '卫衣连帽', '休闲皮鞋', '棒球帽', '围巾', '运动短裤'],
        '食品饮料': ['有机牛奶', '坚果礼盒', '绿茶龙井', '咖啡豆', '巧克力', '蜂蜜柚子茶', '红酒', '蛋白粉', '橄榄油', '燕窝'],
        '家居用品': ['乳胶枕头', '蚕丝被', '空气净化器', '台灯LED', '收纳箱', '四件套', '扫地机器人', '加湿器', '地毯', '毛巾'],
        '运动户外': ['瑜伽垫', '登山包', '帐篷', '跑步腰包', '游泳镜', '羽毛球拍', '头盔', '滑雪手套', '防晒衣', '望远镜'],
    }
    cities = ['北京', '上海', '广州', '深圳', '杭州', '成都', '武汉', '南京']

    rows = []
    for i in range(1, 111):
        cat = random.choice(list(categories.keys()))
        product = random.choice(categories[cat])
        price = round(random.uniform(29, 2999), 2)
        qty = random.randint(1, 4)
        amount = round(price * qty, 2)
        date = datetime(2026, random.randint(1, 8), random.randint(1, 28))
        city = random.choice(cities)
        rows.append([date.strftime('%Y-%m-%d'), product, cat, city, qty, price, amount])

    for i, row_data in enumerate(rows, 2):
        for j, val in enumerate(row_data, 1):
            ws.cell(row=i, column=j, value=val)
    style_data(ws, 2, len(rows) + 1, len(headers))

    path = os.path.join(OUTPUT_DIR, 'day06-透视表基础.xlsx')
    wb.save(path)
    print(f'  [OK] {os.path.basename(path)}')


def generate_day09():
    """Day 9: 统计函数"""
    wb = Workbook()
    add_instruction_sheet(wb, 'Day 9 练习：统计函数', [
        '使用"销售记录"工作表的数据，完成以下函数练习。在数据右边的空白列写入公式：',
        '',
        '【基础统计】',
        '1. =SUM(金额列) → 总销售额',
        '2. =AVERAGE(金额列) → 平均每单金额',
        '3. =COUNT(金额列) → 有多少笔订单',
        '4. =MAX(金额列) 和 =MIN(金额列) → 最大/最小订单金额',
        '',
        '【条件统计】',
        '5. =COUNTIF(分类列,"电子产品") → 电子产品卖了多少单',
        '6. =SUMIF(分类列,"电子产品",金额列) → 电子产品总销售额',
        '7. =COUNTIFS(分类列,"电子产品",城市列,"北京") → 北京电子产品订单数',
        '8. =SUMIFS(金额列,分类列,"服装鞋帽",金额列,">500") → 服装类且>500的销售额',
        '【备用实战】"靓号库(备用实战)"工作表模拟靓号库，练COUNTIF统计类别(规律类型)用',
        '',
        '【思考】哪个分类的平均客单价最高？用AVERAGEIF试试。（答案见下方折叠区）',
    ], thinking=DAY09_THINKING)

    ws = wb.create_sheet('销售记录')
    headers = ['订单ID', '分类', '城市', '金额', '日期']
    style_header(ws, headers, [12, 14, 10, 12, 14])

    categories = ['电子产品', '服装鞋帽', '食品饮料', '家居用品', '运动户外']
    cities = ['北京', '上海', '广州', '深圳', '杭州', '成都']

    rows = []
    for i in range(1, 41):
        cat = random.choice(categories)
        city = random.choice(cities)
        amount = round(random.uniform(30, 3000), 2)
        date = datetime(2026, random.randint(1, 8), random.randint(1, 28))
        rows.append([f'ORD-{i:04d}', cat, city, amount, date.strftime('%Y-%m-%d')])

    for i, row_data in enumerate(rows, 2):
        for j, val in enumerate(row_data, 1):
            ws.cell(row=i, column=j, value=val)
    style_data(ws, 2, len(rows) + 1, len(headers))

    # 在右边留出练习区域
    ws.cell(row=1, column=7, value='← 在这里写公式练习').font = Font(name='微软雅黑', size=10, color='888888')
    ws.column_dimensions['G'].width = 30
    ws.column_dimensions['H'].width = 20

    add_data_sheet(wb, '靓号库(备用实战)', ['靓号ID', '位数', '类别', '价格(乐元)', '购买方式', '赠送VIP'], generate_haoid_rows(80))

    path = os.path.join(OUTPUT_DIR, 'day09-统计函数.xlsx')
    wb.save(path)
    print(f'  [OK] {os.path.basename(path)}')


def generate_day10():
    """Day 10: VLOOKUP"""
    wb = Workbook()
    add_instruction_sheet(wb, 'Day 10 练习：VLOOKUP', [
        '本文件有两张工作表：用户表 和 订单表。',
        '订单表里只有"用户ID"，没有用户姓名和等级。你需要用VLOOKUP从用户表匹配过来。',
        '',
        '【练习】',
        '1. 在订单表的E列（空白列），用VLOOKUP匹配"用户姓名"：',
        '   =VLOOKUP(B2, 用户表!A:D, 2, FALSE)',
        '2. 在F列匹配"用户等级"：',
        '   =VLOOKUP(B2, 用户表!A:D, 3, FALSE)',
        '3. 在G列匹配"城市"：',
        '   =VLOOKUP(B2, 用户表!A:D, 4, FALSE)',
        '',
        '【进阶】',
        '4. 试试如果写成 =VLOOKUP(B2, 用户表!A:D, 2, TRUE) 会发生什么？（近似匹配 vs 精确匹配）',
        '5. 故意把某个订单的用户ID改成一个不存在的ID（如U9999），看看VLOOKUP返回什么（#N/A）',
        '【备用实战】"活动参与名单(备用实战)"工作表的用户ID与"用户表"对齐，直接VLOOKUP练匹配用户等级',
        '',
        '【注意】VLOOKUP只能从左往右查。如果想"根据姓名查ID"（从右往左），需要用INDEX+MATCH或XLOOKUP（XLOOKUP需Excel 2021/365或WPS 2021以上版本）。',
    ], thinking=DAY10_THINKING)

    # 用户表
    ws_users = wb.create_sheet('用户表')
    user_headers = ['用户ID', '用户姓名', '等级', '城市']
    style_header(ws_users, user_headers, [12, 14, 10, 10])

    users = [
        ('U1001', '张三', '金牌', '北京'),
        ('U1002', '李四', '银牌', '上海'),
        ('U1003', '王五', '铜牌', '广州'),
        ('U1004', '赵六', '金牌', '深圳'),
        ('U1005', '孙七', '银牌', '杭州'),
        ('U1006', '周八', '铜牌', '成都'),
        ('U1007', '吴九', '金牌', '武汉'),
        ('U1008', '郑十', '银牌', '南京'),
        ('U1009', '钱十一', '铜牌', '北京'),
        ('U1010', '冯十二', '金牌', '上海'),
        ('U1011', '陈十三', '银牌', '广州'),
        ('U1012', '褚十四', '铜牌', '深圳'),
    ]
    for i, u in enumerate(users, 2):
        for j, val in enumerate(u, 1):
            ws_users.cell(row=i, column=j, value=val)
    style_data(ws_users, 2, len(users) + 1, len(user_headers))

    # 订单表
    ws_orders = wb.create_sheet('订单表')
    order_headers = ['订单ID', '用户ID', '金额', '日期', '用户姓名（VLOOKUP）', '等级（VLOOKUP）', '城市（VLOOKUP）']
    style_header(ws_orders, order_headers, [12, 12, 12, 14, 22, 18, 18])

    orders = []
    for i in range(1, 26):
        user = random.choice(users)
        amount = round(random.uniform(50, 2000), 2)
        date = datetime(2026, random.randint(1, 8), random.randint(1, 28))
        orders.append([f'ORD-{i:04d}', user[0], amount, date.strftime('%Y-%m-%d'), '', '', ''])

    for i, row_data in enumerate(orders, 2):
        for j, val in enumerate(row_data, 1):
            ws_orders.cell(row=i, column=j, value=val)
    style_data(ws_orders, 2, len(orders) + 1, len(order_headers))

    # 说明表保持在最前,打开文件先看到练习说明(VLOOKUP不受工作表顺序影响)
    add_data_sheet(wb, '活动参与名单(备用实战)', ['用户ID', '昵称', '参与活动', '参与日期'], generate_participant_rows(25))

    path = os.path.join(OUTPUT_DIR, 'day10-VLOOKUP.xlsx')
    wb.save(path)
    print(f'  [OK] {os.path.basename(path)}')


def generate_day11():
    """Day 11: 逻辑与文本函数"""
    wb = Workbook()
    add_instruction_sheet(wb, 'Day 11 练习：逻辑与文本函数', [
        '本文件包含商品编码数据，请完成以下练习：',
        '',
        '【IF 逻辑判断】',
        '1. 在E列（空白列），用IF判断价格档次：',
        '   =IF(D2>1000,"高价",IF(D2>300,"中价","低价"))',
        '2. 试试用IFS（Excel 2016+）：=IFS(D2>1000,"高价",D2>300,"中价",TRUE,"低价")',
        '',
        '【文本截取 LEFT/RIGHT/MID】',
        '3. 在F列，用LEFT提取SKU的前3位（产品类别）：=LEFT(B2,3)',
        '4. 在G列，用MID提取SKU的第5-9位（品牌代码）：=MID(B2,5,5)',
        '5. 在H列，用RIGHT提取SKU的最后3位（序号）：=RIGHT(B2,3)',
        '',
        '【TEXT 格式化】',
        '6. 在I列，把日期列格式化为中文：=TEXT(C2,"yyyy年mm月dd日")',
        '7. 在J列，显示星期几：=TEXT(C2,"aaaa")（中文版Excel/WPS用aaaa；英文版Excel/WPS用"dddd"）',
        '【备用实战】"靓号规律判断(备用实战)"工作表只给号码，用函数在空白列判断类别(规律类型)',
        '',
        '【综合】试着用IF+LEFT组合：如果SKU前3位是"ELC"就显示"电子产品"，否则显示"其他"',
    ], thinking=DAY11_THINKING)

    ws = wb.create_sheet('商品编码')
    headers = ['商品名', 'SKU编码', '上架日期', '价格', '价格档次(IF)', '类别(LEFT)', '品牌(MID)', '序号(RIGHT)', '日期格式化(TEXT)', '星期(TEXT)']
    style_header(ws, headers, [16, 18, 14, 10, 16, 14, 14, 14, 22, 14])

    products = [
        ('iPhone 15', 'ELC-APPLE-001', '2026-01-15', 1299),
        ('运动跑鞋', 'SPT-NIKEC-002', '2026-02-20', 599),
        ('有机牛奶', 'FOD-MENGN-003', '2026-03-10', 12.9),
        ('乳胶枕头', 'HOM-LATEX-004', '2026-04-05', 299),
        ('登山包', 'OUT-OSPRE-005', '2026-05-18', 459),
        ('MacBook Pro', 'ELC-APPLE-006', '2026-06-22', 2499),
        ('羽绒服', 'CLO-BOSID-007', '2026-07-01', 899),
        ('坚果礼盒', 'FOD-THREE-008', '2026-08-06', 168),
        ('扫地机器人', 'HOM-ROBOR-009', '2026-01-20', 1299),
        ('望远镜', 'OUT-NIKON-010', '2026-02-14', 299),
        ('华为Mate 60', 'ELC-HUWEI-011', '2026-03-25', 899),
        ('纯棉T恤', 'CLO-UNIQL-012', '2026-04-18', 39),
        ('进口红酒', 'FOD-PENFO-013', '2026-05-30', 258),
        ('空气净化器', 'HOM-XIAOM-014', '2026-06-15', 1999),
        ('跑步腰包', 'OUT-DECAT-015', '2026-07-22', 89),
    ]
    for i, (name, sku, date_str, price) in enumerate(products, 2):
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value=sku)
        ws.cell(row=i, column=3, value=date_str)
        ws.cell(row=i, column=4, value=price)
    style_data(ws, 2, len(products) + 1, len(headers))

    ws_judge = wb.create_sheet('靓号规律判断(备用实战)')
    style_header(ws_judge, ['靓号ID', '类别(用函数判断)'], [18, 24])
    haoid_nums = [r[0] for r in generate_haoid_rows(30)]
    for i, num in enumerate(haoid_nums, 2):
        ws_judge.cell(row=i, column=1, value=num)
    style_data(ws_judge, 2, len(haoid_nums) + 1, 2)

    path = os.path.join(OUTPUT_DIR, 'day11-逻辑与文本函数.xlsx')
    wb.save(path)
    print(f'  [OK] {os.path.basename(path)}')


def generate_day12():
    """Day 12: 图表"""
    wb = Workbook()
    add_instruction_sheet(wb, 'Day 12 练习：图表', [
        '本文件包含两张工作表，用于练习三种图表类型：',
        '',
        '【月度统计】→ 柱状图 + 折线图',
        '1. 选中"月份"和"销售额"两列→插入→柱状图→各月销售额对比',
        '2. 选中"月份"和"订单数"两列→插入→折线图→月度订单趋势',
        '3. 组合图：选中月份+销售额+订单数→插入→组合图→柱子=销售额，折线=订单数',
        '   （右键图表→更改图表类型→组合。WPS更直接：插入→全部图表→组合图，量级差太多记得给折线勾"次坐标轴"）',
        '',
        '【分类统计】→ 饼图',
        '4. 选中"分类"和"销售额"两列→插入→饼图→看各分类销售占比',
        '5. 给饼图加上数据标签（显示百分比）',
        '',
        '【美化】给每张图加上标题、数据标签，调整颜色。',
    ], thinking=DAY12_THINKING)

    # 月度统计
    ws1 = wb.create_sheet('月度统计')
    headers1 = ['月份', '销售额（万元）', '订单数', '客单价（元）']
    style_header(ws1, headers1, [14, 18, 14, 16])

    monthly = [
        ('1月', 85.6, 2140, 400),
        ('2月', 62.3, 1558, 400),
        ('3月', 98.7, 2320, 425),
        ('4月', 105.2, 2391, 440),
        ('5月', 92.1, 2193, 420),
        ('6月', 118.5, 2633, 450),
        ('7月', 132.8, 2951, 450),
        ('8月', 125.4, 2726, 460),
        ('9月', 108.9, 2475, 440),
        ('10月', 145.3, 3125, 465),
        ('11月', 168.7, 3515, 480),
        ('12月', 189.2, 3784, 500),
    ]
    for i, row_data in enumerate(monthly, 2):
        for j, val in enumerate(row_data, 1):
            ws1.cell(row=i, column=j, value=val)
    style_data(ws1, 2, len(monthly) + 1, len(headers1))

    # 分类统计
    ws2 = wb.create_sheet('分类统计')
    headers2 = ['分类', '销售额（万元）', '占比']
    style_header(ws2, headers2, [16, 18, 12])

    cat_data = [
        ('电子产品', 420.5, ''),
        ('服装鞋帽', 288.3, ''),
        ('家居用品', 215.7, ''),
        ('食品饮料', 186.2, ''),
        ('运动户外', 158.6, ''),
    ]
    total = sum(r[1] for r in cat_data)
    for i, (cat, sales, _) in enumerate(cat_data, 2):
        ws2.cell(row=i, column=1, value=cat)
        ws2.cell(row=i, column=2, value=sales)
        ws2.cell(row=i, column=3, value=f'{sales/total*100:.1f}%')
    style_data(ws2, 2, len(cat_data) + 1, len(headers2))

    path = os.path.join(OUTPUT_DIR, 'day12-图表.xlsx')
    wb.save(path)
    print(f'  [OK] {os.path.basename(path)}')


def generate_sql_cheatsheet():
    """SQL速查表（Day 19-31）"""
    wb = Workbook()

    # Sheet1: SQLZoo表结构
    ws1 = wb.active
    ws1.title = 'SQLZoo表结构'
    ws1.column_dimensions['A'].width = 16
    ws1.column_dimensions['B'].width = 16
    ws1.column_dimensions['C'].width = 50

    ws1.cell(row=1, column=1, value='SQLZoo 表结构速查').font = Font(name='微软雅黑', bold=True, size=14)

    tables = [
        ('world表（World Facts）', [
            ('name', 'VARCHAR(50)', '国家名称'),
            ('continent', 'VARCHAR(20)', '大洲（Asia/Europe/Africa...）'),
            ('area', 'INT', '面积（平方公里）'),
            ('population', 'BIGINT', '人口'),
            ('gdp', 'BIGINT', 'GDP（美元）'),
            ('capital', 'VARCHAR(50)', '首都名称'),
        ]),
        ('game表（UEFA Euro）', [
            ('id', 'INT', '比赛ID'),
            ('mdate', 'DATE', '比赛日期'),
            ('stadium', 'VARCHAR(50)', '体育场'),
            ('team1', 'VARCHAR(50)', '队伍1'),
            ('team2', 'VARCHAR(50)', '队伍2'),
        ]),
        ('goal表（UEFA Euro）', [
            ('matchid', 'INT', '比赛ID（关联game.id）'),
            ('teamid', 'VARCHAR(10)', '队伍ID'),
            ('player', 'VARCHAR(50)', '进球球员'),
            ('gtime', 'INT', '进球时间（分钟）'),
        ]),
    ]

    row = 3
    for table_name, cols in tables:
        ws1.cell(row=row, column=1, value=table_name).font = Font(name='微软雅黑', bold=True, size=11, color='4472C4')
        row += 1
        ws1.cell(row=row, column=1, value='列名').font = Font(name='微软雅黑', bold=True, size=10)
        ws1.cell(row=row, column=2, value='类型').font = Font(name='微软雅黑', bold=True, size=10)
        ws1.cell(row=row, column=3, value='说明').font = Font(name='微软雅黑', bold=True, size=10)
        row += 1
        for col_name, col_type, desc in cols:
            ws1.cell(row=row, column=1, value=col_name).font = Font(name='微软雅黑', size=10)
            ws1.cell(row=row, column=2, value=col_type).font = Font(name='微软雅黑', size=10)
            ws1.cell(row=row, column=3, value=desc).font = Font(name='微软雅黑', size=10)
            row += 1
        row += 1

    # Sheet2: 语法速查
    ws2 = wb.create_sheet('语法速查')
    ws2.column_dimensions['A'].width = 22
    ws2.column_dimensions['B'].width = 55
    ws2.column_dimensions['C'].width = 35

    ws2.cell(row=1, column=1, value='SQL 语法速查表').font = Font(name='微软雅黑', bold=True, size=14)
    ws2.cell(row=2, column=1, value='Day 19-31 自学参考').font = Font(name='微软雅黑', size=10, color='888888')

    syntax = [
        ('基础查询', [
            ('SELECT 列名 FROM 表名', '选取指定列的所有行', 'SELECT name, population FROM world'),
            ('SELECT * FROM 表名', '选取所有列', 'SELECT * FROM world'),
            ('SELECT DISTINCT 列名', '去重查询', 'SELECT DISTINCT continent FROM world'),
        ]),
        ('条件筛选', [
            ('WHERE 条件', '筛选符合条件的行', 'WHERE population > 100000000'),
            ('AND / OR', '多条件组合', 'WHERE area > 100000 AND population > 10000000'),
            ('IN (值1, 值2)', '等于其中任意一个', 'WHERE continent IN ("Asia","Europe")'),
            ('BETWEEN A AND B', '在范围内（含边界）', 'WHERE population BETWEEN 1000000 AND 10000000'),
            ('LIKE "模式"', '模糊匹配，%任意字符，_单个字符', 'WHERE name LIKE "A%"'),
        ]),
        ('排序与限制', [
            ('ORDER BY 列名 ASC/DESC', '排序', 'ORDER BY population DESC'),
            ('LIMIT N', '只返回前N行', 'LIMIT 10'),
            ('OFFSET N', '跳过前N行', 'LIMIT 10 OFFSET 20'),
        ]),
        ('聚合函数', [
            ('COUNT(*)', '计数', 'SELECT COUNT(*) FROM world'),
            ('SUM(列)', '求和', 'SELECT SUM(population) FROM world'),
            ('AVG(列)', '平均值', 'SELECT AVG(population) FROM world'),
            ('MAX(列) / MIN(列)', '最大/最小值', 'SELECT MAX(population) FROM world'),
        ]),
        ('分组', [
            ('GROUP BY 列名', '按列分组', 'SELECT continent, COUNT(*) FROM world GROUP BY continent'),
            ('HAVING 条件', '分组后筛选', 'HAVING COUNT(*) > 10'),
        ]),
        ('连接', [
            ('INNER JOIN 表 ON 条件', '内连接（两表都匹配的）', 'FROM game JOIN goal ON game.id=goal.matchid'),
            ('LEFT JOIN 表 ON 条件', '左连接（保留左表所有行）', 'FROM users LEFT JOIN orders ON users.id=orders.uid'),
            ('自连接（SELF JOIN）', '表自己连自己', 'FROM emp e JOIN emp m ON e.mgr_id=m.id'),
        ]),
        ('子查询', [
            ('WHERE 列 IN (SELECT...)', '子查询在WHERE中', 'WHERE population > (SELECT population FROM world WHERE name="Russia")'),
            ('FROM (SELECT...) AS 别名', '子查询在FROM中', 'FROM (SELECT continent, COUNT(*) AS n FROM world GROUP BY continent) AS tmp'),
        ]),
        ('窗口函数', [
            ('ROW_NUMBER() OVER(...)', '分组编号', 'ROW_NUMBER() OVER (PARTITION BY continent ORDER BY population DESC)'),
            ('LAG(列,1) OVER(...)', '取上一行的值', 'LAG(sales,1) OVER (ORDER BY date)'),
            ('LEAD(列,1) OVER(...)', '取下一行的值', 'LEAD(sales,1) OVER (ORDER BY date)'),
        ]),
    ]

    row = 4
    for section, items in syntax:
        ws2.cell(row=row, column=1, value=section).font = Font(name='微软雅黑', bold=True, size=11, color='4472C4')
        row += 1
        for syntax_text, desc, example in items:
            ws2.cell(row=row, column=1, value=syntax_text).font = Font(name='Consolas', size=10)
            ws2.cell(row=row, column=2, value=desc).font = Font(name='微软雅黑', size=10)
            ws2.cell(row=row, column=3, value=example).font = Font(name='Consolas', size=10)
            for c in range(1, 4):
                ws2.cell(row=row, column=c).border = THIN_BORDER
            row += 1
        row += 1

    # Sheet3: 执行顺序
    ws3 = wb.create_sheet('SQL执行顺序')
    ws3.column_dimensions['A'].width = 10
    ws3.column_dimensions['B'].width = 20
    ws3.column_dimensions['C'].width = 55

    ws3.cell(row=1, column=1, value='SQL 语句执行顺序').font = Font(name='微软雅黑', bold=True, size=14)
    ws3.cell(row=2, column=1, value='面试高频考点：书写顺序 ≠ 执行顺序').font = Font(name='微软雅黑', size=10, color='888888')

    headers = ['步骤', '关键字', '作用']
    style_header(ws3, headers, [10, 20, 55])

    order = [
        ('1', 'FROM', '确定数据来源（从哪张表取）。如果有JOIN，先构建虚拟大表'),
        ('2', 'WHERE', '对原始行进行筛选，不满足条件的行被丢弃'),
        ('3', 'GROUP BY', '按指定列分组，每组内的行被"压缩"'),
        ('4', 'HAVING', '对分组后的结果进行筛选（类似WHERE但作用于分组后）'),
        ('5', 'SELECT', '选取要展示的列，计算聚合函数和表达式'),
        ('6', 'ORDER BY', '对最终结果排序（这时SELECT已执行，可以用别名）'),
        ('7', 'LIMIT / OFFSET', '限制返回行数（最后一步）'),
    ]
    for i, (step, keyword, desc) in enumerate(order, 3):
        ws3.cell(row=i, column=1, value=step).font = Font(name='微软雅黑', bold=True, size=10)
        ws3.cell(row=i, column=2, value=keyword).font = Font(name='Consolas', bold=True, size=10, color='4472C4')
        ws3.cell(row=i, column=3, value=desc).font = Font(name='微软雅黑', size=10)
        for c in range(1, 4):
            ws3.cell(row=i, column=c).border = THIN_BORDER

    ws3.cell(row=12, column=1, value='记忆口诀：FWGH SOL — From Where Group Having Select Order Limit').font = Font(name='微软雅黑', size=10, color='E74C3C')

    add_thinking_sheet(wb, SQL_THINKING)

    path = os.path.join(OUTPUT_DIR, 'day19-31-SQL速查表.xlsx')
    wb.save(path)
    print(f'  [OK] {os.path.basename(path)}')


def generate_day19_reference():
    """Day 19 SQL入门参考：一个小型CSV数据集供理解表结构"""
    wb = Workbook()

    ws = wb.active
    ws.title = '示例数据-world风格'

    ws.cell(row=1, column=1, value='示例数据：模拟 world 表（理解表结构用）').font = Font(name='微软雅黑', bold=True, size=12)
    ws.cell(row=2, column=1, value='这份数据模拟了SQLZoo的world表。你可以在脑海中对它"写SQL"，然后对照SQLZoo验证。').font = Font(name='微软雅黑', size=10, color='666666')

    headers = ['name', 'continent', 'area', 'population', 'gdp']
    style_header(ws, headers, [16, 16, 14, 16, 18])

    countries = [
        ('China', 'Asia', 9596961, 1425893465, 17963200000000),
        ('India', 'Asia', 3287263, 1428627663, 3416000000000),
        ('United States', 'North America', 9833520, 339996563, 26950000000000),
        ('Indonesia', 'Asia', 1904569, 277534122, 1319000000000),
        ('Brazil', 'South America', 8515767, 216422446, 1920000000000),
        ('Russia', 'Europe', 17098246, 144444359, 2240000000000),
        ('Japan', 'Asia', 377975, 123294513, 4320000000000),
        ('Germany', 'Europe', 357114, 83294633, 4072000000000),
        ('United Kingdom', 'Europe', 243610, 67736802, 3144000000000),
        ('France', 'Europe', 640679, 64756584, 2958000000000),
        ('Canada', 'North America', 9984670, 38781291, 2168000000000),
        ('Australia', 'Oceania', 7692024, 26439111, 1693000000000),
        ('South Korea', 'Asia', 100210, 51784059, 1798000000000),
        ('Nigeria', 'Africa', 923768, 223804632, 477000000000),
        ('Egypt', 'Africa', 1002450, 112716598, 469000000000),
        ('Mexico', 'North America', 1964375, 128455567, 1677000000000),
        ('Argentina', 'South America', 2780400, 45773884, 631000000000),
        ('South Africa', 'Africa', 1221037, 60414495, 419000000000),
        ('Thailand', 'Asia', 513120, 71801279, 522000000000),
        ('Vietnam', 'Asia', 331212, 98186856, 433000000000),
    ]
    for i, country in enumerate(countries, 4):
        for j, val in enumerate(country, 1):
            ws.cell(row=i, column=j, value=val)
    style_data(ws, 4, 3 + len(countries), len(headers))

    # 练习提示
    tip_row = 4 + len(countries) + 1
    ws.cell(row=tip_row, column=1, value='试着在脑子里写SQL：').font = Font(name='微软雅黑', bold=True, size=10)
    tips = [
        'SELECT name, population FROM [这张表] WHERE continent = "Asia"',
        'SELECT continent, COUNT(*) FROM [这张表] GROUP BY continent',
        'SELECT name FROM [这张表] WHERE population > 100000000 AND area > 1000000',
        'SELECT continent, SUM(population) FROM [这张表] GROUP BY continent ORDER BY SUM(population) DESC',
    ]
    for i, tip in enumerate(tips):
        ws.cell(row=tip_row + 1 + i, column=1, value=tip).font = Font(name='Consolas', size=10, color='4472C4')

    add_thinking_sheet(wb, DAY19_THINKING)

    path = os.path.join(OUTPUT_DIR, 'day19-SQL入门-示例数据.xlsx')
    wb.save(path)
    print(f'  [OK] {os.path.basename(path)}')


# ==================== 备用实战数据(没有真实数据时的替身) ====================
# 结构、字段、价格分布与用户真实业务数据一致
HAOID_PRICE_RANGES = [(10, 49, 15), (50, 99, 35), (100, 180, 35), (181, 500, 15)]  # 引流/主力/利润/展示

def generate_haoid_rows(n=100):
    """模拟靓号库: 列名与真实后台一致(靓号ID/位数/类别/价格/购买方式/赠送VIP)"""
    patterns = ['888', '顺子', '豹子', 'AABB', '普通']
    weights = [20, 8, 5, 12, 55]
    buy_methods = ['直接购买', '竞拍']
    vip_gifts = ['无', 'VIP 1个月', 'VIP 3个月', 'VIP 6个月']
    rows = []
    for _ in range(n):
        digits = random.randint(5, 8)
        pattern = random.choices(patterns, weights=weights)[0]
        # 首位1-9,避免前导0(CSV被Excel打开时前导0会被丢掉)
        num = str(random.randint(1, 9)) + ''.join(str(random.randint(0, 9)) for _ in range(digits - 1))
        if pattern == '888' and digits >= 4:
            num = num[:-3] + '888'
        elif pattern == '顺子':
            start = random.randint(0, 6)
            seq = ''.join(str((start + k) % 10) for k in range(min(4, digits)))
            num = num[:-len(seq)] + seq
        elif pattern == '豹子' and digits >= 5:
            d = str(random.randint(0, 9))
            num = num[:-4] + d * 4
        elif pattern == 'AABB' and digits >= 5:
            d1, d2 = str(random.randint(0, 9)), str(random.randint(0, 9))
            num = num[:-4] + d1 + d1 + d2 + d2
        lo, hi, _ = random.choices(HAOID_PRICE_RANGES, weights=[r[2] for r in HAOID_PRICE_RANGES])[0]
        price = round(random.uniform(lo, hi))
        status = random.choices(['在售', '已售'], weights=[65, 35])[0]
        if status == '在售':
            method, vip = '未售', '无'
        else:
            method = random.choice(buy_methods)
            vip = random.choice(vip_gifts)
        rows.append([num, digits, pattern, price, method, vip])
    return rows

def generate_activity_rows():
    """模拟13场活动: 活动名/类型/日期/参与人数/购买人数/发券数/转化率"""
    events = [
        ('春节福袋', '节日', '2026-02-10'), ('520告白季', '折扣', '2026-05-20'),
        ('双11狂欢', '折扣', '2025-11-11'), ('圣诞竞拍夜', '竞拍', '2025-12-24'),
        ('端午打卡', '打卡', '2026-06-19'), ('元旦集卡', '打卡', '2026-01-01'),
        ('暑期福利日', '折扣', '2026-07-15'), ('中秋特卖', '节日', '2025-10-06'),
        ('国庆竞拍', '竞拍', '2025-10-01'), ('五一特惠', '折扣', '2026-05-01'),
        ('元宵灯会', '节日', '2026-03-05'), ('清明踏青', '打卡', '2026-04-05'),
        ('618年中庆', '折扣', '2026-06-18'),
    ]
    rows = []
    for name, typ, date in events:
        participants = random.randint(80, 600)
        buyers = random.randint(int(participants * 0.2), int(participants * 0.6))
        coupons = random.randint(int(participants * 0.5), participants)
        conv = f'{buyers / participants * 100:.1f}%'
        rows.append([name, typ, date, participants, buyers, coupons, conv])
    return rows

def generate_participant_rows(n=30):
    """模拟活动参与名单: 用户ID/昵称/参与活动/参与日期(含重复用户,供去重和VLOOKUP练习)"""
    nicknames = ['小鱼', '阿明', '柠檬茶', '夜风', '青柠', '北岛', '星星', '大树', '麦子', '浮云',
                 '海盐', '奶茶', '柚子', '白鹿', '山川', '风筝', '月亮', '夏夜', '落落', '石头']
    events = ['春节福袋', '520告白季', '双11狂欢', '圣诞竞拍夜', '端午打卡', '元旦集卡', '618年中庆']
    user_ids = [f'U{1001 + i}' for i in range(12)]  # 与day10用户表ID对齐,VLOOKUP可直用
    rows = []
    for _ in range(n):
        uid = random.choice(user_ids)
        nick = random.choice(nicknames)
        event = random.choice(events)
        date = f'2026-{random.randint(1, 8):02d}-{random.randint(1, 28):02d}'
        rows.append([uid, nick, event, date])
    # 故意加几条完全重复的行(供去重练习)
    for _ in range(4):
        rows.append(rows[random.randint(0, n - 1)][:])
    return rows

def generate_activity_detail_rows(n=300):
    """模拟活动参与明细(阶段三AI辅助分析用): 用户ID/活动名/类型/日期/是否购买/消费金额"""
    events = [('春节福袋', '节日'), ('520告白季', '折扣'), ('双11狂欢', '折扣'),
              ('圣诞竞拍夜', '竞拍'), ('端午打卡', '打卡'), ('元旦集卡', '打卡'),
              ('暑期福利日', '折扣'), ('618年中庆', '折扣')]
    user_ids = [str(48360000 + i) for i in range(80)]
    rows = []
    for _ in range(n):
        uid = random.choice(user_ids)
        name, typ = random.choice(events)
        date = f'2026-{random.randint(1, 8):02d}-{random.randint(1, 28):02d}'
        bought = random.choices([1, 0], weights=[35, 65])[0]
        amount = round(random.uniform(50, 180)) if bought else 0
        rows.append([uid, name, typ, date, bought, amount])
    return rows


def generate_backup_csvs():
    """备用数据CSV(阶段三AI辅助分析+阶段四项目备用)"""
    import csv as csv_mod
    haoid = [['靓号ID', '位数', '类别(规律类型)', '价格(乐元)', '购买方式', '赠送VIP', '上架日期']]
    for row in generate_haoid_rows(120):
        haoid.append(row[:6] + [f'2026-{random.randint(1, 8):02d}-{random.randint(1, 28):02d}'])
    with open(os.path.join(OUTPUT_DIR, '备用数据-靓号库.csv'), 'w', newline='', encoding='utf-8-sig') as f:
        csv_mod.writer(f).writerows(haoid)
    print('  [OK] 备用数据-靓号库.csv')

    detail = [['用户ID', '活动名', '活动类型', '参与日期', '是否购买', '消费金额(乐元)']]
    detail += generate_activity_detail_rows(300)
    with open(os.path.join(OUTPUT_DIR, '备用数据-活动明细.csv'), 'w', newline='', encoding='utf-8-sig') as f:
        csv_mod.writer(f).writerows(detail)
    print('  [OK] 备用数据-活动明细.csv')


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    print('正在生成练习文件...\n')

    generate_day02()
    generate_day03()
    generate_day04()
    generate_day05()
    generate_day06()  # Day 6-7 共用
    generate_day09()
    generate_day10()
    generate_day11()
    generate_day12()
    generate_day19_reference()
    generate_sql_cheatsheet()
    generate_backup_csvs()

    print(f'\n[OK] 全部完成！文件保存在 {OUTPUT_DIR}')
    print(f'共 {len(os.listdir(OUTPUT_DIR))} 个文件')


if __name__ == '__main__':
    main()
